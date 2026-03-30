# excelembeds/excelembeds.py
import asyncio
import io
import json
import logging
import re
from datetime import datetime, timezone, timedelta
from typing import Any, Dict, List, Optional

import discord
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from redbot.core import checks, commands, Config
from redbot.core.bot import Red
from redbot.core.utils import bounded_gather


class Excelembeds(commands.Cog):
    """Excelembeds – Excel → Rich Embeds (2026 polished edition)."""

    MAX_FILE_SIZE_MB = 5
    DEFAULT_REMINDER_MINUTES = [60, 30, 15, 5]
    DEFAULT_REMINDER_EMOJI = "🔔"

    def __init__(self, bot: Red):
        self.bot = bot
        self.logger = logging.getLogger("red.excelembeds")
        self.config = Config.get_conf(self, identifier=987654321987654321, force_registration=True)
        defaults_guild = {
            "reminder_mode": False,
            "reminder_minutes": self.DEFAULT_REMINDER_MINUTES,
            "max_rows": 50,
            "pending_reminders": {},
        }
        self.config.register_guild(**defaults_guild)
        self.reminder_task: Optional[asyncio.Task] = None

    async def cog_load(self):
        self.reminder_task = asyncio.create_task(self._reminder_loop())
        self.logger.info("Excelembeds cog loaded successfully.")

    async def cog_unload(self):
        if self.reminder_task and not self.reminder_task.done():
            self.reminder_task.cancel()
            try:
                await self.reminder_task
            except asyncio.CancelledError:
                pass
        self.logger.info("Excelembeds cog unloaded.")

    async def red_delete_data_for_user(self, *, requester: str, user_id: int):
        """Delete user reminder data when requested (Red GDPR hook)."""
        guilds = await self.config.all_guilds()
        for guild_id, data in guilds.items():
            pending = data.get("pending_reminders", {})
            modified = False
            for msg_id, rem in list(pending.items()):
                if user_id in rem.get("users", []):
                    rem["users"] = [u for u in rem["users"] if u != user_id]
                    modified = True
            if modified:
                await self.config.guild_from_id(guild_id).pending_reminders.set(pending)
        self.logger.info(f"Deleted reminder data for user {user_id} (requested by {requester})")

    async def _reminder_loop(self):
        await self.bot.wait_until_red_ready()
        while True:
            try:
                await asyncio.sleep(120)
                all_guilds = await self.config.all_guilds()
                for guild_id, data in all_guilds.items():
                    if not data.get("reminder_mode"):
                        continue
                    guild = self.bot.get_guild(guild_id)
                    if not guild:
                        continue
                    pending = data.get("pending_reminders", {}).copy()
                    if not pending:
                        continue
                    now = datetime.now(timezone.utc)
                    modified = False
                    for msg_id_str, rem in list(pending.items()):
                        try:
                            event_time = datetime.fromisoformat(rem["event_time"])
                        except Exception:
                            pending.pop(msg_id_str, None)
                            modified = True
                            continue
                        if now > event_time + timedelta(hours=6):
                            pending.pop(msg_id_str, None)
                            modified = True
                            continue
                        users = rem.get("users", [])
                        sent = rem.get("sent", {})
                        changed = False
                        intervals = rem.get("reminder_minutes") or data.get("reminder_minutes", self.DEFAULT_REMINDER_MINUTES)
                        for interval in intervals:
                            reminder_time = event_time - timedelta(minutes=interval)
                            if now >= reminder_time and str(interval) not in sent:
                                tasks = [
                                    self._send_dm_reminder(guild.get_member(uid), event_time, interval)
                                    for uid in users
                                    if guild.get_member(uid) and str(uid) not in sent.get(str(interval), [])
                                ]
                                if tasks and len(tasks) > 25 and guild.owner:
                                    try:
                                        await guild.owner.send(f"⚠️ **Excelembeds rate-limit warning**: Sending {len(tasks)} DMs for message {msg_id_str}.")
                                    except Exception:
                                        pass
                                await bounded_gather(*tasks, return_exceptions=True)
                                sent[str(interval)] = [uid for uid in users]
                                changed = True
                        if changed:
                            rem["sent"] = sent
                            modified = True
                    if modified:
                        await self.config.guild(guild).pending_reminders.set(pending)
            except asyncio.CancelledError:
                break
            except Exception as e:
                self.logger.exception("Reminder loop error")
                await asyncio.sleep(30)

    async def _send_dm_reminder(self, member: Optional[discord.Member], event_time: datetime, minutes_before: int):
        if not member:
            return
        try:
            embed = discord.Embed(
                title="🔔 Event Reminder",
                description=f"Your event is in **{minutes_before} minutes**!\n{event_time.strftime('%A, %B %d at %I:%M %p %Z')}",
                color=discord.Color.gold(),
            )
            await asyncio.wait_for(member.send(embed=embed), timeout=8)
            await asyncio.sleep(1.0)
        except Exception:
            pass

    def _normalize_key(self, name: str) -> str:
        return str(name).strip().lower().replace(" ", "").replace("_", "")

    def _get_column_indices(self, headers: List[str]) -> Dict[str, int]:
        aliases = {
            "content": ["content", "message", "text"],
            "title": ["title", "embedtitle"],
            "description": ["description", "desc"],
            "color": ["color", "colour", "embedcolor"],
            "url": ["url", "titleurl"],
            "image": ["image", "embedimage", "imageurl"],
            "thumbnail": ["thumbnail", "thumb"],
            "author_name": ["authorname", "author"],
            "author_url": ["authorurl"],
            "author_icon": ["authoricon"],
            "footer_text": ["footer", "footertext"],
            "footer_icon": ["footericon"],
            "timestamp": ["timestamp", "time"],
            "fields": ["fields", "embedfields"],
            "buttons": ["buttons", "embedbuttons"],
            "dropdowns": ["dropdowns", "selects", "multiselect"],
            "event_time": ["eventtime", "starttime", "datetime", "eventdate"],
            "ping_role": ["pingrole", "roleid", "mentionrole"],
            "reminder_minutes": ["reminderminutes", "reminders", "remindertimes"],
            "reminder_emoji": ["reminderemoji", "reminderreaction"],
        }
        col_map: Dict[str, int] = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            norm = self._normalize_key(h)
            for canonical, alias_list in aliases.items():
                if any(norm == self._normalize_key(a) for a in alias_list):
                    col_map[canonical] = i
                    break
            else:
                col_map[norm] = i
        return col_map

    def _get_cell(self, row: tuple, col_map: Dict[str, int], key: str, default: Any = None) -> Any:
        idx = col_map.get(key)
        return row[idx] if idx is not None and idx < len(row) else default

    def _parse_datetime(self, value: Any) -> Optional[datetime]:
        if not value:
            return None
        if isinstance(value, (int, float)):
            try:
                return (datetime(1899, 12, 30) + timedelta(days=value)).replace(tzinfo=timezone.utc)
            except Exception:
                pass
        value_str = str(value).strip()
        if not value_str:
            return None
        formats = ["%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%m/%d/%Y %H:%M", "%m/%d/%Y %H:%M:%S",
                   "%m/%d/%y %H:%M", "%m/%d/%y %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%d/%m/%Y %H:%M"]
        for fmt in formats:
            try:
                return datetime.strptime(value_str, fmt).replace(tzinfo=timezone.utc)
            except ValueError:
                continue
        return None

    def _parse_color(self, color_val: Any) -> Optional[discord.Color]:
        if not color_val:
            return None
        color_str = str(color_val).strip().lower()
        try:
            return discord.Color(int(color_str[1:], 16)) if color_str.startswith("#") else discord.Color.from_str(color_str)
        except Exception:
            return None

    def _format_mentions(self, text: str, guild: discord.Guild) -> str:
        if not text:
            return text
        text = re.sub(r"(?<!<@&)(\d{17,19})(?!>)", lambda m: f"<@&{m.group(0)}>" if guild.get_role(int(m.group(0))) else m.group(0), text)
        text = re.sub(r"(?<!<#)(\d{17,19})(?!>)", lambda m: f"<#{m.group(0)}>" if guild.get_channel(int(m.group(0))) else m.group(0), text)
        return text

    def _validate_image_url(self, url: str) -> bool:
        if not url or not url.startswith(("http://", "https://")):
            return False
        return any(url.lower().endswith(ext) for ext in (".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"))

    def _build_embed_from_row(self, row: tuple, col_map: Dict[str, int], guild: discord.Guild) -> Optional[discord.Embed]:
        title = str(self._get_cell(row, col_map, "title", "")).strip()[:256]
        if not title and not self._get_cell(row, col_map, "description"):
            return None

        embed = discord.Embed(
            title=title or None,
            description=str(self._get_cell(row, col_map, "description", "")).strip()[:4096] or None,
            color=self._parse_color(self._get_cell(row, col_map, "color")),
            url=str(self._get_cell(row, col_map, "url", "")).strip() or None,
            timestamp=self._parse_datetime(self._get_cell(row, col_map, "timestamp")),
        )

        image = str(self._get_cell(row, col_map, "image", "")).strip()
        if image and self._validate_image_url(image):
            embed.set_image(url=image)
        thumbnail = str(self._get_cell(row, col_map, "thumbnail", "")).strip()
        if thumbnail and self._validate_image_url(thumbnail):
            embed.set_thumbnail(url=thumbnail)

        author_name = str(self._get_cell(row, col_map, "author_name", "")).strip()[:256]
        if author_name:
            embed.set_author(
                name=author_name,
                url=str(self._get_cell(row, col_map, "author_url", "")).strip() or None,
                icon_url=str(self._get_cell(row, col_map, "author_icon", "")).strip() or None,
            )
        footer_text = str(self._get_cell(row, col_map, "footer_text", "")).strip()[:2048]
        if footer_text:
            embed.set_footer(
                text=footer_text,
                icon_url=str(self._get_cell(row, col_map, "footer_icon", "")).strip() or None,
            )

        fields_json = self._get_cell(row, col_map, "fields")
        if fields_json:
            try:
                fields_list = json.loads(str(fields_json).strip())
                if isinstance(fields_list, list):
                    for f in fields_list[:25]:
                        if isinstance(f, dict):
                            embed.add_field(
                                name=str(f.get("name", ""))[:256],
                                value=str(f.get("value", ""))[:1024],
                                inline=bool(f.get("inline", False)),
                            )
            except Exception:
                pass
        return embed

    def _build_view_from_row(self, row: tuple, col_map: Dict[str, int]) -> Optional[discord.ui.View]:
        buttons_json = self._get_cell(row, col_map, "buttons")
        dropdowns_json = self._get_cell(row, col_map, "dropdowns")

        class DynamicView(discord.ui.View):
            def __init__(self, timeout: Optional[float] = None):
                super().__init__(timeout=timeout)

            async def _generic_callback(self, interaction: discord.Interaction):
                await interaction.response.send_message("✅ Interaction received!", ephemeral=True)

        view = DynamicView(timeout=None)

        if buttons_json:
            try:
                btn_list = json.loads(str(buttons_json).strip())
                if isinstance(btn_list, list):
                    for btn_data in btn_list[:25]:
                        if not isinstance(btn_data, dict):
                            continue
                        label = str(btn_data.get("label", "Button"))[:80]
                        url = str(btn_data.get("url", "")).strip()
                        style_str = str(btn_data.get("style", "primary")).lower()
                        style_map = {"primary": 1, "secondary": 2, "success": 3, "danger": 4, "link": 5}
                        style = style_map.get(style_str, 1)
                        if url:
                            style = 5
                        btn = discord.ui.Button(
                            label=label,
                            url=url or None,
                            style=discord.ButtonStyle(style),
                            emoji=btn_data.get("emoji"),
                            row=int(btn_data.get("row", 0)) % 5,
                            disabled=bool(btn_data.get("disabled", False)),
                            custom_id=f"excelembeds:btn:{label[:20]}" if not url else None,
                        )
                        if not url:
                            btn.callback = view._generic_callback
                        view.add_item(btn)
            except Exception:
                pass

        dropdown_list = []
        if dropdowns_json:
            try:
                dropdown_list = json.loads(str(dropdowns_json).strip())
                if not isinstance(dropdown_list, list):
                    dropdown_list = [dropdown_list]
            except Exception:
                dropdown_list = []
        for dd_data in dropdown_list[:5]:
            if not isinstance(dd_data, dict):
                continue
            options = [discord.SelectOption(label=str(opt)[:100]) for opt in dd_data.get("options", [])[:25]]
            if not options:
                continue
            select = discord.ui.Select(
                placeholder=str(dd_data.get("placeholder", "Select..."))[:150],
                options=options,
                min_values=int(dd_data.get("min_values", 1)),
                max_values=int(dd_data.get("max_values", 1)),
                disabled=bool(dd_data.get("disabled", False)),
                row=int(dd_data.get("row", 0)) % 5,
                custom_id=f"excelembeds:select:{dd_data.get('placeholder', '')[:20]}",
            )
            select.callback = view._generic_callback
            view.add_item(select)

        return view if view.children else None

    @commands.Cog.listener()
    async def on_interaction(self, interaction: discord.Interaction):
        if not interaction.custom_id or not interaction.custom_id.startswith("excelembeds:"):
            return
        try:
            await interaction.response.send_message("✅ Interaction received!", ephemeral=True)
        except Exception:
            pass

    @commands.group(name="excelembed", aliases=["xlembed", "excelembeds"], invoke_without_command=True)
    @commands.guild_only()
    @checks.admin_or_permissions(manage_messages=True)
    @commands.bot_has_permissions(send_messages=True, embed_links=True)
    async def excelembed(self, ctx: commands.Context):
        """Excelembeds – Excel to rich embeds.

        Type ,excelembeds by itself to see this help menu and all subcommands.
        """
        if ctx.invoked_subcommand is None:
            await ctx.send_help()

    @excelembed.command(name="guide")
    async def excelembed_guide(self, ctx: commands.Context):
        """Compact guide – all headers, formatting, optional columns, and config commands."""
        embed = discord.Embed(
            title="Excelembeds Guide",
            color=discord.Color.blue(),
            description="Upload `.xlsx` → rich embeds + buttons/dropdowns + reminders.\n"
                        "**Get started:** `[p]excelembed template`",
        )

        embed.add_field(
            name="Core Headers",
            value=(
                "`title` / `description` — **at least one required**\n"
                "`content` — text above embed (optional)\n"
                "`color` — #hex or name (optional)\n"
                "`url` — title hyperlink (optional)"
            ),
            inline=False,
        )
        embed.add_field(
            name="Media & Author",
            value=(
                "`image` / `thumbnail` — direct image URL (optional)\n"
                "`author_name` / `author_url` / `author_icon` — author block (optional)\n"
                "`footer_text` / `footer_icon` — footer (optional)"
            ),
            inline=False,
        )
        embed.add_field(
            name="Advanced",
            value=(
                "`timestamp` — date/time (optional)\n"
                "`fields` — JSON list of `{\"name\":.., \"value\":.., \"inline\":bool}` (optional)\n"
                "`buttons` — JSON list `{\"label\":.., \"url\":.., \"style\":primary/link, \"emoji\":.., \"row\":0}` (optional)\n"
                "`dropdowns` — JSON list of selects `{\"placeholder\":.., \"options\":[\"A\",\"B\"], \"min_values\":1, \"max_values\":1, \"row\":0}` (optional)"
            ),
            inline=False,
        )
        embed.add_field(
            name="Mentions & Reminders",
            value=(
                "`ping_role` — role ID (auto <@&ID>)\n"
                "`event_time` — date/time for reminders (optional)\n"
                "`reminder_minutes` — JSON list e.g. `[60,30,15]` (overrides guild default)\n"
                "`reminder_emoji` — custom reaction emoji (default 🔔)"
            ),
            inline=False,
        )
        embed.add_field(
            name="Commands",
            value=(
                "`create #channel [yes]` — send all rows\n"
                "`preview #channel 5` — test row 5 only\n"
                "`template` — download example file"
            ),
            inline=False,
        )
        embed.add_field(
            name="Config (`[p]excelembed config`)",
            value=(
                "`maxrows <number>` — set max rows per file (1-200)\n"
                "`reminders` — toggle DM reminders on/off\n"
                "`cleanup` — clear all pending reminders"
            ),
            inline=False,
        )
        embed.set_footer(text="One row = one embed | JSON columns must be valid JSON")
        await ctx.send(embed=embed)

    @excelembed.command(name="template")
    async def excelembed_template(self, ctx: commands.Context):
        """Send the ready-to-use Excel template."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Embed Template"
        ws.append([
            "content", "title", "description", "color", "url", "image", "thumbnail",
            "author_name", "author_url", "author_icon", "footer_text", "footer_icon",
            "timestamp", "fields", "buttons", "dropdowns", "event_time", "ping_role",
            "reminder_minutes", "reminder_emoji"
        ])
        ws.append([
            "Announcement!", "Community Event",
            "Join us! <@&123456789> in <#987654321>",
            "#00FF00", "https://example.com", "https://i.imgur.com/example.png", "",
            "Event Host", "", "https://i.imgur.com/host.png",
            "Powered by Excelembeds", "",
            "2026-04-15 19:00",
            '[{"name":"Date","value":"April 15","inline":true}]',
            '[{"label":"RSVP","url":"https://example.com","emoji":"✅","style":"primary","row":0}]',
            '[{"placeholder":"Choose role","options":["Member","VIP"],"min_values":1,"max_values":1,"row":1}]',
            "2026-04-15 19:00",
            "123456789",
            '[60,30,15]',
            "🔔"
        ])
        ws.append(["← Fill rows below. One row = one embed."])
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        file = discord.File(buffer, filename="excelembed_template.xlsx")
        await ctx.send("**Excelembeds Template** – Attach to `[p]excelembed create` or `[p]excelembed preview`", file=file)

    @excelembed.command(name="preview")
    async def excelembed_preview(self, ctx: commands.Context, channel: Optional[discord.TextChannel] = None, row_number: int = 1):
        """Preview a single row from the attached Excel file (no reminders)."""
        if not ctx.message.attachments:
            return await ctx.send("❌ Attach an `.xlsx` file.")
        attachment = ctx.message.attachments[0]
        if attachment.size > self.MAX_FILE_SIZE_MB * 1024 * 1024:
            return await ctx.send(f"❌ File too large (max {self.MAX_FILE_SIZE_MB} MB).")
        if not attachment.filename.lower().endswith((".xlsx", ".xls")):
            return await ctx.send("❌ Only `.xlsx` files supported.")

        try:
            data = await attachment.read()
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
            ws = wb.active
        except Exception as e:
            return await ctx.send(f"❌ Invalid Excel: {str(e)[:200]}")

        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        col_map = self._get_column_indices(headers)

        row_data = None
        for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if r_idx == row_number:
                row_data = row
                break
        if not row_data:
            return await ctx.send(f"❌ Row {row_number} not found.")

        try:
            embed = self._build_embed_from_row(row_data, col_map, ctx.guild)
            if not embed:
                return await ctx.send("❌ Invalid embed (no title/description).")

            content = str(self._get_cell(row_data, col_map, "content", "")).strip()[:2000] or None
            ping_role_id = str(self._get_cell(row_data, col_map, "ping_role", "")).strip()
            if ping_role_id.isdigit() and ctx.guild.get_role(int(ping_role_id)):
                content = f"<@&{ping_role_id}> {content or ''}".strip()

            if content:
                content = self._format_mentions(content, ctx.guild)
            if embed.description:
                embed.description = self._format_mentions(embed.description, ctx.guild)
            for field in embed.fields:
                field.value = self._format_mentions(field.value, ctx.guild)

            view = self._build_view_from_row(row_data, col_map)
            target = channel or ctx.channel
            await target.send(content=content, embed=embed, view=view)
            await ctx.send(f"✅ Preview of row **{row_number}** sent to {target.mention}.")
        except Exception as exc:
            await ctx.send(f"❌ Preview error: {str(exc)[:200]}")

    @excelembed.command(name="create", aliases=["send"])
    async def excelembed_create(self, ctx: commands.Context, channel: discord.TextChannel, reminders: str = "no"):
        """Import Excel and send rich embeds."""
        if not ctx.message.attachments:
            return await ctx.send("❌ Attach an `.xlsx` file.")

        attachment = ctx.message.attachments[0]
        if attachment.size > self.MAX_FILE_SIZE_MB * 1024 * 1024:
            return await ctx.send(f"❌ File too large (max {self.MAX_FILE_SIZE_MB} MB).")
        if not attachment.filename.lower().endswith((".xlsx", ".xls")):
            return await ctx.send("❌ Only `.xlsx` files supported.")

        try:
            data = await attachment.read()
        except Exception:
            return await ctx.send("❌ Failed to download file.")

        reminders_enabled = reminders.lower() in ("yes", "true", "on", "1")
        if reminders_enabled:
            await self.config.guild(ctx.guild).reminder_mode.set(True)

        try:
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
            ws = wb.active
        except Exception as e:
            return await ctx.send(f"❌ Invalid Excel: {str(e)[:200]}")

        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        col_map = self._get_column_indices(headers)
        max_rows = await self.config.guild(ctx.guild).max_rows()

        rows_processed = 0
        errors = []

        if len(list(ws.iter_rows(min_row=2, values_only=True))) > max_rows:
            await ctx.send(f"⚠️ Warning: File has more than {max_rows} rows – only first {max_rows} will be processed.")

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if rows_processed >= max_rows:
                break
            if all(v is None for v in row):
                continue

            try:
                embed = self._build_embed_from_row(row, col_map, ctx.guild)
                if not embed:
                    errors.append(f"Row {row_idx}: Skipped (invalid embed).")
                    continue

                content = str(self._get_cell(row, col_map, "content", "")).strip()[:2000] or None
                ping_role_id = str(self._get_cell(row, col_map, "ping_role", "")).strip()
                if ping_role_id.isdigit() and ctx.guild.get_role(int(ping_role_id)):
                    content = f"<@&{ping_role_id}> {content or ''}".strip()

                if content:
                    content = self._format_mentions(content, ctx.guild)
                if embed.description:
                    embed.description = self._format_mentions(embed.description, ctx.guild)
                for field in embed.fields:
                    field.value = self._format_mentions(field.value, ctx.guild)

                view = self._build_view_from_row(row, col_map)
                message = await channel.send(content=content, embed=embed, view=view)

                event_time = self._parse_datetime(self._get_cell(row, col_map, "event_time"))
                if reminders_enabled and event_time:
                    emoji = str(self._get_cell(row, col_map, "reminder_emoji", self.DEFAULT_REMINDER_EMOJI)).strip() or self.DEFAULT_REMINDER_EMOJI
                    await message.add_reaction(emoji)
                    pending = await self.config.guild(ctx.guild).pending_reminders()
                    pending[str(message.id)] = {
                        "event_time": event_time.isoformat(),
                        "users": [],
                        "sent": {},
                        "reminder_minutes": json.loads(str(self._get_cell(row, col_map, "reminder_minutes", "[]")).strip()) or None,
                        "emoji": emoji,
                    }
                    await self.config.guild(ctx.guild).pending_reminders.set(pending)

                rows_processed += 1
            except Exception as exc:
                errors.append(f"Row {row_idx}: Error – {str(exc)[:150]}")

        msg = f"✅ **Success!** Sent **{rows_processed}** embed(s) to {channel.mention}."
        if errors:
            msg += "\n\n**Warnings/Errors:**\n" + "\n".join(errors[:15])
        if reminders_enabled and ctx.guild.member_count > 500:
            msg += "\n\n⚠️ **Large guild warning**: DM reminders may hit rate limits."
        await ctx.send(msg)

    @commands.Cog.listener()
    async def on_raw_reaction_add(self, payload: discord.RawReactionActionEvent):
        guild = self.bot.get_guild(payload.guild_id)
        if not guild or not await self.config.guild(guild).reminder_mode():
            return
        pending = await self.config.guild(guild).pending_reminders()
        key = str(payload.message_id)
        if key not in pending:
            return
        rem = pending[key]
        emoji = rem.get("emoji", self.DEFAULT_REMINDER_EMOJI)
        if str(payload.emoji) != emoji or payload.user_id == self.bot.user.id:
            return
        if payload.user_id not in rem["users"]:
            rem["users"].append(payload.user_id)
            await self.config.guild(guild).pending_reminders.set(pending)

    @excelembed.group(name="config")
    @checks.admin_or_permissions(manage_guild=True)
    async def excelembed_config(self, ctx: commands.Context):
        """Guild configuration for Excelembeds."""
        pass

    @excelembed_config.command(name="maxrows")
    async def config_maxrows(self, ctx: commands.Context, number: int):
        """Set maximum rows processed per Excel file (default 50)."""
        if number < 1 or number > 200:
            return await ctx.send("❌ Value must be between 1 and 200.")
        await self.config.guild(ctx.guild).max_rows.set(number)
        await ctx.send(f"✅ Max rows per file set to **{number}**.")

    @excelembed_config.command(name="reminders")
    async def config_reminders(self, ctx: commands.Context):
        """Toggle reminder mode and view settings."""
        current = await self.config.guild(ctx.guild).reminder_mode()
        await self.config.guild(ctx.guild).reminder_mode.set(not current)
        conf = await self.config.guild(ctx.guild).all()
        await ctx.send(
            f"**Excelembeds Settings**\n"
            f"Reminder mode: {'✅ Enabled' if conf['reminder_mode'] else '❌ Disabled'}\n"
            f"Default minutes: {conf['reminder_minutes']}\n"
            f"Max rows: {conf['max_rows']}\n"
            f"Active reminders: {len(conf.get('pending_reminders', {}))}"
        )

    @excelembed_config.command(name="cleanup")
    async def config_cleanup(self, ctx: commands.Context):
        """Clear all pending reminders."""
        await self.config.guild(ctx.guild).pending_reminders.set({})
        await ctx.send("✅ All pending reminders cleared.")
