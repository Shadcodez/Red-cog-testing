# ExcelEvents/core.py
import asyncio
import aiohttp
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional, Dict, List, Tuple

import discord
import openpyxl
from redbot.core import Config, commands
from redbot.core.bot import Red

from .utils import _parse_datetime, _get_column_indices, _get_cell, _normalize_key, _is_valid_xlsx


class ExcelEvents(commands.Cog):
    """Bulk Discord Scheduled Events from Excel/CSV with refined image support."""

    MAX_ROWS = 500
    MAX_IMAGE_SIZE = 15 * 1024 * 1024

    def __init__(self, bot: Red):
        self.bot = bot
        self.config = Config.get_conf(self, identifier=9876543210987654321, force_registration=True)
        defaults_guild = {
            "event_mappings": {},
            "last_synced": None,
            "announcement_mode": False,
            "announcement_channel": None,
            "reminder_mode": False,
            "reminder_channel": None,
            "reminder_minutes": [60, 15, 5],
            "reminder_sent": {},
        }
        self.config.register_guild(**defaults_guild)
        self.reminder_task = None
        self.session = None

        # Attach commands
        from .commands import attach_commands
        attach_commands(self)

    async def cog_load(self):
        self.session = aiohttp.ClientSession()
        if self.reminder_task is None or self.reminder_task.done():
            self.reminder_task = asyncio.create_task(self._reminder_task())

    def cog_unload(self):
        if self.session:
            asyncio.create_task(self.session.close())
        if self.reminder_task and not self.reminder_task.done():
            self.reminder_task.cancel()

    # ====================== REFINED IMAGE DOWNLOADER ======================
    async def _download_image(self, url: str) -> Optional[bytes]:
        """Final refined image downloader with retries and strong browser headers."""
        if not url or not str(url).startswith(("http://", "https://")):
            return None

        url = url.strip()

        # Imgur fixes
        if "imgur.com" in url:
            url = url.replace(".jpeg", ".jpg").replace(".JPEG", ".jpg")
            if "i.imgur.com" not in url and "imgur.com" in url:
                image_id = url.split("/")[-1].split("?")[0].split(".")[0]
                url = f"https://i.imgur.com/{image_id}.jpg"
            if "?" in url:
                url = url.split("?")[0]

        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                          "(KHTML, like Gecko) Chrome/134.0.0.0 Safari/537.36",
            "Accept": "image/avif,image/webp,image/apng,image/*,*/*;q=0.8",
            "Referer": "https://imgur.com/",
        }

        for attempt in range(2):
            try:
                async with self.session.get(url, headers=headers, timeout=25, allow_redirects=True) as resp:
                    if resp.status != 200:
                        continue
                    content_length = int(resp.headers.get("Content-Length", 0))
                    if content_length > self.MAX_IMAGE_SIZE or content_length == 0:
                        continue
                    data = await resp.read()
                    content_type = resp.headers.get("Content-Type", "").lower()
                    if len(data) > 10240 and (
                        any(x in content_type for x in ("image/", "jpeg", "png", "gif", "webp")) or
                        data.startswith((b'\xff\xd8', b'\x89PNG', b'GIF8'))
                    ):
                        return data
            except Exception:
                if attempt == 0:
                    await asyncio.sleep(1.5)
                continue
        return None

    async def _create_event_with_image(self, guild: discord.Guild, data: Dict, image_bytes: Optional[bytes] = None) -> Optional[discord.ScheduledEvent]:
        name = str(data.get("name", "")).strip()
        if not name or len(name) > 100:
            return None

        start_time = await _parse_datetime(data.get("start"))
        if not start_time:
            return None

        end_time = await _parse_datetime(data.get("end"))
        description = str(data.get("description", "")).strip()[:1000] or None
        event_type_str = str(data.get("type", "")).strip().lower() or "voice"
        location = str(data.get("location", "")).strip() or None
        channel_id_input = data.get("channelid")

        if event_type_str in ("external", "url", "link"):
            entity_type = discord.EntityType.external
        elif event_type_str == "stage":
            entity_type = discord.EntityType.stage_instance
        else:
            entity_type = discord.EntityType.voice

        channel = None
        if entity_type in (discord.EntityType.voice, discord.EntityType.stage_instance) and channel_id_input:
            try:
                ch_id = int(str(channel_id_input).strip())
                temp_ch = guild.get_channel(ch_id)
                if temp_ch and (
                    (entity_type == discord.EntityType.voice and isinstance(temp_ch, discord.VoiceChannel)) or
                    (entity_type == discord.EntityType.stage_instance and isinstance(temp_ch, discord.StageChannel))
                ):
                    channel = temp_ch
            except Exception:
                pass

        try:
            if entity_type == discord.EntityType.external:
                if not location:
                    return None
                event = await guild.create_scheduled_event(
                    name=name, description=description, start_time=start_time,
                    end_time=end_time, entity_type=entity_type, location=location,
                    privacy_level=discord.PrivacyLevel.guild_only
                )
            else:
                if not channel:
                    return None
                event = await guild.create_scheduled_event(
                    name=name, description=description, start_time=start_time,
                    end_time=end_time, entity_type=entity_type, channel=channel,
                    privacy_level=discord.PrivacyLevel.guild_only
                )

            if image_bytes:
                try:
                    await event.edit(cover=image_bytes)
                except Exception:
                    pass  # image failed but event still created

            await asyncio.sleep(1.8)
            return event
        except Exception:
            return None

    async def _update_event(self, event: discord.ScheduledEvent, data: Dict, image_bytes: Optional[bytes] = None):
        try:
            edit_kwargs = {
                "name": str(data.get("name", "")).strip(),
                "description": str(data.get("description", "")).strip()[:1000] or None,
                "start_time": await _parse_datetime(data.get("start")),
                "end_time": await _parse_datetime(data.get("end")),
            }
            if image_bytes:
                edit_kwargs["cover"] = image_bytes
            await event.edit(**edit_kwargs)
            await asyncio.sleep(1.2)
        except Exception:
            pass

    def _create_event_embed(self, event: discord.ScheduledEvent) -> discord.Embed:
        embed = discord.Embed(
            title=event.name[:256],
            description=(event.description or "No description provided.")[:4096],
            color=discord.Color.blurple(),
            url=event.url,
        )
        if event.start_time:
            ts = int(event.start_time.timestamp())
            embed.add_field(name="🕒 Start", value=f"<t:{ts}:F> (<t:{ts}:R>)", inline=True)
        if event.end_time:
            ts = int(event.end_time.timestamp())
            embed.add_field(name="🕒 End", value=f"<t:{ts}:F>", inline=True)

        loc = event.location or (event.channel.mention if event.channel else "Voice/Stage")
        embed.add_field(name="📍 Location", value=loc, inline=False)
        embed.add_field(name="Type", value=event.entity_type.name.replace("_", " ").title(), inline=True)
        embed.set_footer(text="New Event • Synced via ExcelEvents • RedBot 2026")
        return embed

    def _create_reminder_embed(self, event: discord.ScheduledEvent, minutes: int) -> discord.Embed:
        embed = discord.Embed(
            title=f"⏰ {event.name} starts in {minutes} minutes!",
            description=(event.description or "")[:4096],
            color=discord.Color.orange(),
            url=event.url,
        )
        if event.start_time:
            ts = int(event.start_time.timestamp())
            embed.add_field(name="Exact Time", value=f"<t:{ts}:F>", inline=False)
        loc = event.location or (event.channel.mention if event.channel else "Voice/Stage")
        embed.add_field(name="📍 Location", value=loc, inline=False)
        embed.set_footer(text=f"Reminder • ExcelEvents")
        return embed

    async def _validate_excel(self, file_path: Path, guild: discord.Guild) -> Tuple[List[str], List[str]]:
        errors: List[str] = []
        warnings: List[str] = []

        if not file_path.exists():
            errors.append("No events.xlsx file found. Use `upload` or `paste` first.")
            return errors, warnings

        if file_path.stat().st_size == 0:
            errors.append("The uploaded file is empty.")
            return errors, warnings

        is_real_xlsx = _is_valid_xlsx(file_path)

        try:
            if not is_real_xlsx:
                raise zipfile.BadZipFile("Not a valid .xlsx")

            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            if ws is None or ws.max_row < 1:
                errors.append("Worksheet is empty or unreadable.")
                return errors, warnings

            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [str(cell).strip().lower() if cell is not None else "" for cell in header_row]
            col_map = _get_column_indices(headers)

        except (zipfile.BadZipFile, openpyxl.utils.exceptions.InvalidFileException):
            errors.append("❌ This is **not** a valid .xlsx file. Use `paste` instead.")
            return errors, warnings
        except Exception as e:
            errors.append(f"Failed to read file: {type(e).__name__} – {e}")
            return errors, warnings

        required = {"name", "start"}
        missing = [col for col in required if col not in col_map]
        if missing:
            errors.append(f"Missing required column(s): {', '.join(missing)}")

        row_num = 1
        seen_names: set[str] = set()

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_num += 1
            if not row or all(v is None for v in row):
                continue

            name = str(_get_cell(row, col_map, "name", "")).strip()
            if not name:
                errors.append(f"Row {row_num}: Missing or empty **Name**")
                continue

            if len(name) > 100:
                errors.append(f"Row {row_num}: Name too long (max 100 characters)")

            key = _normalize_key(name)
            if key in seen_names:
                warnings.append(f"Row {row_num}: Duplicate name '{name}'")
            seen_names.add(key)

            start_dt = await _parse_datetime(_get_cell(row, col_map, "start"))
            if not start_dt:
                errors.append(f"Row {row_num}: Invalid **Start** time format")
            elif start_dt < datetime.now(timezone.utc):
                warnings.append(f"Row {row_num}: Start time is in the past")

        if not seen_names:
            errors.append("No valid event rows found in the file.")

        return errors, warnings

    async def _reminder_task(self):
        await self.bot.wait_until_ready()
        while True:
            try:
                for guild in self.bot.guilds:
                    config = self.config.guild(guild)
                    if not await config.reminder_mode():
                        continue
                    ch_id = await config.reminder_channel()
                    channel = guild.get_channel(ch_id) if ch_id else None
                    if not (channel and channel.permissions_for(guild.me).send_messages):
                        continue

                    mappings = await config.event_mappings()
                    reminder_sent = await config.reminder_sent() or {}

                    for event_id in list(mappings.values()):
                        try:
                            event = await guild.fetch_scheduled_event(event_id)
                            if event.status not in (discord.ScheduledEventStatus.scheduled, discord.ScheduledEventStatus.active):
                                continue
                            if not event.start_time:
                                continue

                            minutes_until = (event.start_time - datetime.now(timezone.utc)).total_seconds() / 60
                            for min_before in await config.reminder_minutes():
                                if min_before > 0 and abs(minutes_until - min_before) <= 7:
                                    sent_list = reminder_sent.get(str(event_id), [])
                                    if min_before not in sent_list:
                                        embed = self._create_reminder_embed(event, min_before)
                                        await channel.send(embed=embed)
                                        reminder_sent.setdefault(str(event_id), []).append(min_before)
                                        await asyncio.sleep(1.5)
                        except Exception:
                            continue

                    await config.reminder_sent.set(reminder_sent)
            except Exception:
                pass
            await asyncio.sleep(300)
