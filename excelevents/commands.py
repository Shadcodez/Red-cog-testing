# ExcelEvents/commands.py
import asyncio
import io
import csv
from pathlib import Path
from typing import Dict

import discord
import openpyxl
from redbot.core import commands
from redbot.core import data_manager

from .core import ExcelEvents
from .utils import _get_column_indices, _get_cell, _normalize_key, _parse_datetime


def attach_commands(cog: ExcelEvents):

    @cog.excelevents.command(name="guide")
    async def guide(self, ctx: commands.Context):
        """Shows a detailed guide on how to use the ExcelEvents cog."""
        embed = discord.Embed(
            title="📖 ExcelEvents - Complete Guide",
            description="Bulk create Discord Scheduled Events with refined image support.",
            color=discord.Color.blurple()
        )
        embed.add_field(name="Image Tips", value="Use direct links like `https://i.imgur.com/XXXXXX.jpg`\nYou can also attach one image to the `sync` command as fallback.", inline=False)
        embed.add_field(name="Quick Start", value="`template` → fill data → `upload` or `paste` → `check` → `sync`", inline=False)
        await ctx.send(embed=embed)

    @cog.excelevents.command(name="template")
    async def template(self, ctx: commands.Context):
        """Sends a ready-to-use CSV template with image column."""
        example = (
            "name,start,end,description,type,location,channelid,image\n"
            'Game Night,2026-04-05 20:00,2026-04-05 22:00,Weekly game night,voice,,"123456789012345678",https://i.imgur.com/3eQczTs.jpg\n'
        )
        await ctx.send(f"**CSV Template:**\n```csv\n{example}\n```")

    @cog.excelevents.command(name="upload")
    async def upload(self, ctx: commands.Context):
        """Upload an .xlsx file containing events."""
        if not ctx.message.attachments:
            await ctx.send("❌ Please attach an `.xlsx` or `.xls` file.")
            return
        attachment = ctx.message.attachments[0]
        if not attachment.filename.lower().endswith((".xlsx", ".xls")):
            await ctx.send("❌ Only `.xlsx` or `.xls` files are supported.")
            return

        data_path: Path = data_manager.cog_data_path(self)
        data_path.mkdir(parents=True, exist_ok=True)
        file_path = data_path / "events.xlsx"

        if file_path.exists():
            file_path.unlink()

        await attachment.save(str(file_path))
        await ctx.send("✅ File uploaded (old file replaced). Use `check`.")

    @cog.excelevents.command(name="paste")
    async def paste(self, ctx: commands.Context):
        """Paste CSV data to create the events file."""
        lines = ctx.message.content.splitlines()
        csv_text = "\n".join(lines[1:]) if len(lines) > 1 else ""

        if not csv_text.strip():
            await ctx.send("❌ Please paste CSV data after the command.")
            return

        data_path = data_manager.cog_data_path(self)
        data_path.mkdir(parents=True, exist_ok=True)
        file_path = data_path / "events.xlsx"

        if file_path.exists():
            file_path.unlink()

        try:
            input_io = io.StringIO(csv_text.strip())
            reader = csv.reader(input_io, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL, skipinitialspace=True)
            rows = [[cell.strip() for cell in row] for row in reader if row and any(cell.strip() for cell in row)]

            if len(rows) < 1:
                await ctx.send("❌ No valid rows found.")
                return
            if len(rows) - 1 > cog.MAX_ROWS:
                rows = rows[:cog.MAX_ROWS + 1]
                await ctx.send(f"⚠️ Only first {cog.MAX_ROWS} events saved.")

            if rows:
                header_len = len(rows[0])
                for i in range(1, len(rows)):
                    rows[i] += [''] * (header_len - len(rows[i]))

            wb = openpyxl.Workbook()
            ws = wb.active
            for row in rows:
                ws.append(row)
            wb.save(file_path)

            await ctx.send(f"✅ CSV saved! **{len(rows)-1}** events loaded.\nUse `check`.")
        except Exception as e:
            await ctx.send(f"❌ Failed to parse CSV: {type(e).__name__} – {e}")

    @cog.excelevents.command(name="check")
    async def check(self, ctx: commands.Context):
        """Validate the events.xlsx file before syncing."""
        data_path = data_manager.cog_data_path(self)
        file_path = data_path / "events.xlsx"
        await ctx.send("🔍 Running validation...")
        errors, warnings = await self._validate_excel(file_path, ctx.guild)

        if errors:
            await ctx.send("**Validation Failed:**\n" + "\n".join(f"❌ {msg}" for msg in errors))
        elif warnings:
            await ctx.send("**✅ Valid with warnings:**\n" + "\n".join(f"⚠️ {msg}" for msg in warnings) + "\n\nYou may now run `sync`.")
        else:
            await ctx.send("✅ **Perfect!** Ready to sync.")

    @cog.excelevents.command(name="sync")
    async def sync(self, ctx: commands.Context):
        """Sync the spreadsheet → Discord Scheduled Events (create/update/delete + images)."""
        if not ctx.guild.me.guild_permissions.manage_events:
            await ctx.send("❌ I need the **Manage Events** permission.")
            return

        data_path = data_manager.cog_data_path(self)
        file_path = data_path / "events.xlsx"
        if not file_path.exists():
            await ctx.send("❌ No file found. Use `upload` or `paste` first.")
            return

        errors, warnings = await self._validate_excel(file_path, ctx.guild)
        if errors:
            await ctx.send("⚠️ Validation failed. Run `check` first.")
            return

        await ctx.send("🔄 Syncing events with refined image support...")

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [str(cell).strip().lower() if cell is not None else "" for cell in header_row]
            col_map = _get_column_indices(headers)

            global_image_bytes = None
            if ctx.message.attachments:
                att = ctx.message.attachments[0]
                if att.content_type and att.content_type.startswith("image/") and att.size < cog.MAX_IMAGE_SIZE:
                    global_image_bytes = await att.read()

            mappings = await self.config.guild(ctx.guild).event_mappings()
            new_mappings: Dict[str, int] = {}
            active_keys = set()
            processed = 0
            new_events_created = []

            row_num = 1
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_num += 1
                if not row or all(v is None for v in row):
                    continue

                name = str(_get_cell(row, col_map, "name", "")).strip()
                if not name:
                    continue

                key = _normalize_key(name)
                active_keys.add(key)

                data = {
                    "name": name,
                    "start": _get_cell(row, col_map, "start"),
                    "end": _get_cell(row, col_map, "end"),
                    "description": _get_cell(row, col_map, "description"),
                    "type": _get_cell(row, col_map, "type"),
                    "location": _get_cell(row, col_map, "location"),
                    "channelid": _get_cell(row, col_map, "channelid"),
                }

                image_url = str(_get_cell(row, col_map, "image", "")).strip()
                image_bytes = None

                if image_url:
                    image_bytes = await cog._download_image(image_url)
                    if image_bytes:
                        await ctx.send(f"✅ Row {row_num}: Image loaded for **{name}**")
                    else:
                        await ctx.send(f"⚠️ Row {row_num}: Image failed for **{name}** — event created without cover")
                elif global_image_bytes:
                    image_bytes = global_image_bytes
                    await ctx.send(f"✅ Row {row_num}: Using attached image for **{name}**")

                start_time = await _parse_datetime(data["start"])
                if not start_time:
                    continue

                if key in mappings:
                    try:
                        event = await ctx.guild.fetch_scheduled_event(mappings[key])
                        await cog._update_event(event, data, image_bytes)
                        new_mappings[key] = event.id
                        processed += 1
                        continue
                    except Exception:
                        pass

                new_event = await cog._create_event_with_image(ctx.guild, data, image_bytes)
                if new_event:
                    new_mappings[key] = new_event.id
                    new_events_created.append(new_event)
                    processed += 1
                else:
                    await ctx.send(f"⚠️ Failed to create event: {name}")

            # Cleanup
            deleted = 0
            for old_key, old_id in list(mappings.items()):
                if old_key not in active_keys:
                    try:
                        event = await ctx.guild.fetch_scheduled_event(old_id)
                        await event.delete()
                        deleted += 1
                    except Exception:
                        pass

            await self.config.guild(ctx.guild).event_mappings.set(new_mappings)
            await self.config.guild(ctx.guild).last_synced.set(datetime.now(timezone.utc).isoformat())

            # Write IDs and URLs back
            try:
                wb = openpyxl.load_workbook(file_path, data_only=True)
                ws = wb.active
                headers = [str(cell).strip().lower() if cell is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

                id_col = next((c + 1 for c, h in enumerate(headers) if h == "discord event id"), ws.max_column + 1)
                url_col = next((c + 1 for c, h in enumerate(headers) if h == "discord event url"), ws.max_column + 1)

                ws.cell(1, id_col, "Discord Event ID")
                ws.cell(1, url_col, "Discord Event URL")

                for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    name = str(_get_cell(row, col_map, "name", "")).strip()
                    if name and _normalize_key(name) in new_mappings:
                        eid = new_mappings[_normalize_key(name)]
                        try:
                            ev = await ctx.guild.fetch_scheduled_event(eid)
                            ws.cell(r_idx, id_col, eid)
                            ws.cell(r_idx, url_col, ev.url)
                        except Exception:
                            pass
                wb.save(file_path)
            except Exception:
                pass

            # Announcements
            announced = 0
            if await self.config.guild(ctx.guild).announcement_mode():
                ann_ch_id = await self.config.guild(ctx.guild).announcement_channel()
                if ann_ch_id:
                    channel = ctx.guild.get_channel(ann_ch_id)
                    if channel and channel.permissions_for(ctx.guild.me).send_messages:
                        for event in new_events_created:
                            try:
                                await channel.send(embed=self._create_event_embed(event))
                                announced += 1
                                await asyncio.sleep(0.8)
                            except Exception:
                                pass

            result = f"**✅ Sync Complete**\n• Processed: **{processed}**\n• Active: **{len(new_mappings)}**\n• Deleted: **{deleted}**"
            if announced:
                result += f"\n📢 Announced **{announced}** new events!"
            result += "\n📊 Spreadsheet updated with Discord Event IDs & URLs."
            await ctx.send(result)

        except Exception as e:
            await ctx.send(f"❌ Sync failed: {type(e).__name__}: {e}")

    @cog.excelevents.command(name="status")
    async def status(self, ctx: commands.Context):
        """Show current status of the ExcelEvents cog."""
        data_path = data_manager.cog_data_path(self)
        file_path = data_path / "events.xlsx"
        mappings = await self.config.guild(ctx.guild).event_mappings()
        await ctx.send(
            f"**ExcelEvents Status**\n"
            f"• File exists: **{file_path.exists()}**\n"
            f"• Tracked events: **{len(mappings)}**"
        )

    @cog.excelevents.command(name="testimage")
    async def testimage(self, ctx: commands.Context, *, url: str):
        """Debug tool: Test downloading a single image URL."""
        await ctx.send(f"🔍 Testing image: `{url}`")

        image_bytes = await cog._download_image(url)

        if image_bytes:
            size_kb = len(image_bytes) // 1024
            await ctx.send(f"✅ **Success!** Downloaded **{size_kb} KB** image.")
        else:
            await ctx.send("❌ **Failed** to download image.")

        # Extra debug info
        try:
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                              "(KHTML, like Gecko) Chrome/134.0.0.0 Safari/537.36",
            }
            async with cog.session.get(url, headers=headers, timeout=15, allow_redirects=True) as resp:
                await ctx.send(f"Status: **{resp.status}** | Content-Type: `{resp.headers.get('Content-Type', 'None')}`")
        except Exception as e:
            await ctx.send(f"Debug request failed: {type(e).__name__}")

    # Other group commands (announcement, reminder, clear) with descriptions
    @cog.excelevents.group(name="announcement", invoke_without_command=True)
    async def announcement_group(self, ctx: commands.Context):
        """Manage announcement settings for new events."""
        await ctx.send_help(ctx.command)

    @announcement_group.command(name="toggle")
    async def toggle_announcement(self, ctx: commands.Context, channel: Optional[discord.TextChannel] = None):
        """Toggle or set the announcement channel."""
        config = self.config.guild(ctx.guild)
        if channel is None:
            new_mode = not await config.announcement_mode()
            await config.announcement_mode.set(new_mode)
            await ctx.send(f"✅ Announcement mode **{'enabled' if new_mode else 'disabled'}**.")
            return
        await config.announcement_channel.set(channel.id)
        await config.announcement_mode.set(True)
        await ctx.send(f"✅ Announcement mode enabled → {channel.mention}")

    @cog.excelevents.group(name="reminder", invoke_without_command=True)
    async def reminder_group(self, ctx: commands.Context):
        """Manage reminder settings."""
        await ctx.send_help(ctx.command)

    @reminder_group.command(name="toggle")
    async def toggle_reminder(self, ctx: commands.Context, channel: Optional[discord.TextChannel] = None):
        """Toggle or set the reminder channel."""
        config = self.config.guild(ctx.guild)
        if channel is None:
            new_mode = not await config.reminder_mode()
            await config.reminder_mode.set(new_mode)
            await ctx.send(f"✅ Reminder mode **{'enabled' if new_mode else 'disabled'}**.")
            return
        await config.reminder_channel.set(channel.id)
        await config.reminder_mode.set(True)
        await ctx.send(f"✅ Reminder mode enabled → {channel.mention}")

    @reminder_group.command(name="times")
    async def reminder_times(self, ctx: commands.Context, *minutes: int):
        """Set reminder times (in minutes before start)."""
        valid = [m for m in minutes if m > 0]
        if not valid:
            await ctx.send("❌ Please provide positive numbers.")
            return
        await self.config.guild(ctx.guild).reminder_minutes.set(valid)
        await ctx.send(f"✅ Reminder times updated to: **{valid}** minutes before start.")

    @cog.excelevents.command(name="clear")
    async def clear(self, ctx: commands.Context):
        """Delete the events file and reset all mappings."""
        data_path = data_manager.cog_data_path(self)
        file_path = data_path / "events.xlsx"
        if file_path.exists():
            file_path.unlink()
            await self.config.guild(ctx.guild).event_mappings.set({})
            await ctx.send("✅ Events file deleted and mappings reset.")
        else:
            await ctx.send("No file to clear.")


def attach_commands(cog: ExcelEvents):
    # Commands are attached via the decorator pattern in the functions above
    pass
