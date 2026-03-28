import asyncio
import csv
import io
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Optional, Dict, List, Tuple

import discord
import openpyxl
from redbot.core import commands, Config, data_manager
from redbot.core.bot import Red


class ExcelEvents(commands.Cog):
    """Easily create and manage Discord Scheduled Events from Excel or pasted CSV.

    Features:
    • Bulk upload/sync events from .xlsx (or .xls) or raw CSV
    • Forgiving date parser (Excel serial dates + many common formats)
    • Updates existing events or creates new ones (with clear user feedback)
    • Announcement mode: rich embeds posted to a user-set channel when events are created/updated
    • Reminder system: automatic pre-event reminders posted to a separate user-set channel
      (configurable times e.g. 60 min, 15 min before start)
    • Advanced validation (`check` command) with errors + warnings
    • Only ONE Excel file is ever kept (auto-overwrite)
    • Full visibility: every action gives clear output so users always know the cog is working
      and that events were set up properly (with direct links to the Discord events)

    **Quick Start:**
    1. `[p]excelevents upload` → attach your events.xlsx
    2. `[p]excelevents check` → advanced validation
    3. `[p]excelevents sync` → creates/updates events with live feedback + links
    4. `[p]excelevents announcement toggle #announce` → creation announcements
    5. `[p]excelevents reminder toggle #reminders` + `[p]excelevents reminder times 60 15`
    """

    def __init__(self, bot: Red):
        self.bot = bot
        self.config = Config.get_conf(
            self, identifier=9876543210987654321, force_registration=True
        )
        defaults_guild = {
            "event_mappings": {},       # normalized_name → event_id
            "last_synced": None,
            "announcement_mode": False,
            "announcement_channel": None,
            "reminder_mode": False,
            "reminder_channel": None,
            "reminder_minutes": [60, 15],   # minutes before start
            "reminder_sent": {},            # str(event_id) → list of reminded minutes
        }
        self.config.register_guild(**defaults_guild)

        # Background reminder task (persistent across restarts)
        self.reminder_task = None

    async def cog_load(self):
        """Start the reminder background task when the cog loads."""
        if self.reminder_task is None or self.reminder_task.done():
            self.reminder_task = asyncio.create_task(self._reminder_task())

    # ====================== FORGIVING DATE PARSER ======================
    async def _parse_datetime(self, value) -> Optional[datetime]:
        """Parse Excel serial dates, common strings, or return None."""
        if not value:
            return None
        if isinstance(value, (int, float)):
            try:
                base = datetime(1899, 12, 30)
                dt = base + timedelta(days=value)
                return dt.replace(tzinfo=timezone.utc)
            except Exception:
                pass

        value_str = str(value).strip()
        formats = [
            "%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S",
            "%m/%d/%Y %H:%M", "%m/%d/%Y %H:%M:%S",
            "%m/%d/%y %H:%M", "%m/%d/%y %H:%M:%S",
            "%m/%d/%Y %I:%M %p", "%m/%d/%Y %I:%M:%S %p",
            "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %I:%M %p",
        ]
        for fmt in formats:
            try:
                dt = datetime.strptime(value_str, fmt)
                return dt.replace(tzinfo=timezone.utc)
            except ValueError:
                continue
        return None

    def _normalize_key(self, name: str) -> str:
        """Case-insensitive key for event tracking."""
        return str(name).strip().lower()

    async def _create_event(self, guild: discord.Guild, data: Dict) -> Optional[discord.ScheduledEvent]:
        """Create a scheduled event. Returns None on failure."""
        name = str(data.get("name", "")).strip()
        if not name:
            return None

        start_time = await self._parse_datetime(data.get("start"))
        if not start_time:
            return None

        end_time = await self._parse_datetime(data.get("end"))
        description = str(data.get("description", "")).strip()[:1000] or None
        event_type = str(data.get("type", "")).strip().lower() or "voice"
        location = str(data.get("location", "")).strip() or None
        channel_id_input = data.get("channelid")

        channel = None
        event_location = None

        if event_type == "external" and location:
            entity_type = discord.EntityType.external
            event_location = location
        else:
            entity_type = (
                discord.EntityType.stage_instance
                if event_type == "stage"
                else discord.EntityType.voice
            )
            if channel_id_input:
                try:
                    ch_id = int(str(channel_id_input).strip())
                    channel = guild.get_channel(ch_id)
                except Exception:
                    pass

        if entity_type in (discord.EntityType.voice, discord.EntityType.stage_instance) and not channel:
            return None

        try:
            event = await guild.create_scheduled_event(
                name=name,
                description=description,
                start_time=start_time,
                end_time=end_time,
                entity_type=entity_type,
                channel=channel,
                location=event_location,
                privacy_level=discord.PrivacyLevel.guild_only,
            )
            await asyncio.sleep(1.5)
            return event
        except Exception:
            return None

    def _create_event_embed(self, event: discord.ScheduledEvent) -> discord.Embed:
        """Rich announcement embed (used for creation announcements)."""
        embed = discord.Embed(
            title=event.name[:256],
            description=(event.description or "No description provided.")[:4096],
            color=discord.Color.blurple(),
            url=event.url,
        )

        if event.start_time:
            start_ts = int(event.start_time.timestamp())
            embed.add_field(
                name="🕒 Start",
                value=f"<t:{start_ts}:F> (<t:{start_ts}:R>)",
                inline=True,
            )
        if event.end_time:
            end_ts = int(event.end_time.timestamp())
            embed.add_field(
                name="🕒 End",
                value=f"<t:{end_ts}:F>",
                inline=True,
            )

        if event.entity_type == discord.EntityType.external:
            loc_text = event.location or "External link"
        else:
            loc_text = event.channel.mention if event.channel else "Voice/Stage channel"
        embed.add_field(name="📍 Location", value=loc_text, inline=False)

        embed.add_field(
            name="Type",
            value=event.entity_type.name.replace("_", " ").title(),
            inline=True,
        )

        embed.set_footer(text="New Event • Synced via ExcelEvents • RedBot 2026")
        return embed

    def _create_reminder_embed(self, event: discord.ScheduledEvent, minutes: int) -> discord.Embed:
        """Rich reminder embed posted to the reminder channel."""
        embed = discord.Embed(
            title=f"⏰ Event Starting Soon: {event.name}",
            description=(event.description or "No description provided.")[:4096],
            color=discord.Color.orange(),
            url=event.url,
        )

        if event.start_time:
            start_ts = int(event.start_time.timestamp())
            embed.add_field(
                name="Starts in",
                value=f"**{minutes} minutes** (<t:{start_ts}:R>)",
                inline=False,
            )
            embed.add_field(
                name="Exact Time",
                value=f"<t:{start_ts}:F>",
                inline=True,
            )

        if event.entity_type == discord.EntityType.external:
            loc_text = event.location or "External link"
        else:
            loc_text = event.channel.mention if event.channel else "Voice/Stage channel"
        embed.add_field(name="📍 Location", value=loc_text, inline=False)

        embed.add_field(
            name="Type",
            value=event.entity_type.name.replace("_", " ").title(),
            inline=True,
        )

        embed.set_footer(text=f"Reminder • {minutes} min before • ExcelEvents")
        return embed

    # ====================== ADVANCED VALIDATION ======================
    async def _validate_excel(self, file_path: Path, guild: discord.Guild) -> Tuple[List[str], List[str]]:
        """Advanced validation: returns (errors, warnings)."""
        errors: List[str] = []
        warnings: List[str] = []

        if not file_path.exists():
            errors.append("No events.xlsx file found. Use `upload` or `paste` first.")
            return errors, warnings

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            if ws is None:
                errors.append("Could not read the active worksheet.")
                return errors, warnings

            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [str(cell).strip().lower() if cell is not None else "" for cell in header_row]

            required = {"name", "start"}
            missing = [col for col in required if col not in headers]
            if missing:
                errors.append(f"Missing required column(s): {', '.join(missing)}")

            row_num = 1
            seen_names: set[str] = set()

            for row in ws.iter_rows(min_row=2, values_only=True):
                row_num += 1
                if not row or all(v is None for v in row):
                    continue

                data = {
                    headers[i]: row[i]
                    for i in range(len(row))
                    if i < len(headers) and headers[i]
                }

                name = str(data.get("name", "")).strip()
                if not name:
                    errors.append(f"Row {row_num}: Missing or empty **Name**")
                    continue

                if len(name) > 100 or len(name) < 1:
                    errors.append(f"Row {row_num}: Name must be 1-100 characters (currently {len(name)})")

                key = self._normalize_key(name)
                if key in seen_names:
                    warnings.append(f"Row {row_num}: Duplicate event name '{name}' – only the last row will be kept")
                seen_names.add(key)

                event_type = str(data.get("type", "")).strip().lower() or "voice"
                valid_types = {"voice", "stage", "external"}
                if event_type not in valid_types:
                    warnings.append(f"Row {row_num}: Unknown Type '{event_type}' – defaulting to voice")

                start_val = data.get("start")
                start_dt = await self._parse_datetime(start_val)
                if not start_dt:
                    errors.append(f"Row {row_num}: Invalid **Start** time `{start_val}`")
                else:
                    now = datetime.now(timezone.utc)
                    if start_dt < now:
                        warnings.append(f"Row {row_num}: Start time is in the past")

                end_val = data.get("end")
                if end_val is not None and str(end_val).strip():
                    end_dt = await self._parse_datetime(end_val)
                    if not end_dt:
                        errors.append(f"Row {row_num}: Invalid **End** time `{end_val}`")
                    elif start_dt and end_dt <= start_dt:
                        errors.append(f"Row {row_num}: **End** time must be after **Start** time")

                desc = str(data.get("description", "")).strip()
                if len(desc) > 1000:
                    warnings.append(f"Row {row_num}: Description will be truncated to 1000 characters")

                location = str(data.get("location", "")).strip() or None
                channel_id_input = data.get("channelid")

                if event_type == "external":
                    if not location:
                        errors.append(f"Row {row_num}: External events require a **Location**")
                    elif len(location) > 100:
                        errors.append(f"Row {row_num}: External Location too long (max 100 chars)")
                else:
                    if not channel_id_input:
                        warnings.append(f"Row {row_num}: Voice/Stage events should have a **ChannelID**")
                    else:
                        try:
                            ch_id = int(str(channel_id_input).strip())
                            if ch_id <= 0:
                                errors.append(f"Row {row_num}: **ChannelID** must be positive")
                            else:
                                channel = guild.get_channel(ch_id)
                                if not channel:
                                    warnings.append(f"Row {row_num}: ChannelID {ch_id} does not exist")
                                elif event_type == "voice" and channel.type != discord.ChannelType.voice:
                                    warnings.append(f"Row {row_num}: ChannelID {ch_id} is not a Voice channel")
                                elif event_type == "stage" and channel.type != discord.ChannelType.stage_voice:
                                    warnings.append(f"Row {row_num}: ChannelID {ch_id} is not a Stage channel")
                        except ValueError:
                            errors.append(f"Row {row_num}: **ChannelID** must be a valid number")

            if not seen_names:
                errors.append("No valid event rows found in the file.")

            return errors, warnings

        except Exception as e:
            errors.append(f"Failed to read Excel file: {type(e).__name__} – {e}")
            return errors, warnings

    # ====================== BACKGROUND REMINDER TASK ======================
    async def _reminder_task(self):
        """Background task that sends pre-event reminders to the configured channel."""
        await self.bot.wait_until_ready()
        while True:
            try:
                for guild in list(self.bot.guilds):
                    config = self.config.guild(guild)
                    if not await config.reminder_mode():
                        continue

                    ch_id = await config.reminder_channel()
                    if not ch_id:
                        continue
                    channel = guild.get_channel(ch_id)
                    if not (channel and channel.permissions_for(guild.me).send_messages):
                        continue

                    mappings = await config.event_mappings()
                    reminder_sent_raw = await config.reminder_sent() or {}
                    reminder_sent = {k: list(v) for k, v in reminder_sent_raw.items()}
                    updated = False

                    for event_id in mappings.values():
                        event_id_str = str(event_id)
                        try:
                            event = await guild.fetch_scheduled_event(event_id)
                            if event.status not in (discord.ScheduledEventStatus.scheduled, discord.ScheduledEventStatus.active):
                                continue
                            if not event.start_time:
                                continue

                            now = datetime.now(timezone.utc)
                            minutes_until = (event.start_time - now).total_seconds() / 60.0
                            if minutes_until <= 0:
                                continue

                            for min_before in await config.reminder_minutes():
                                if min_before <= 0:
                                    continue
                                # Send if we are within a small window around the target time
                                if abs(minutes_until - min_before) <= 7:
                                    sent_list = reminder_sent.get(event_id_str, [])
                                    if min_before not in sent_list:
                                        embed = self._create_reminder_embed(event, min_before)
                                        await channel.send(embed=embed)
                                        if event_id_str not in reminder_sent:
                                            reminder_sent[event_id_str] = []
                                        reminder_sent[event_id_str].append(min_before)
                                        updated = True
                                        await asyncio.sleep(1.5)
                        except discord.NotFound:
                            continue
                        except Exception:
                            continue

                    if updated:
                        await config.reminder_sent.set(reminder_sent)

            except Exception:
                pass  # Prevent task from crashing

            await asyncio.sleep(300)  # Check every 5 minutes

    # ====================== COMMANDS ======================
    @commands.group(name="excelevents", invoke_without_command=True)
    @commands.guild_only()
    @commands.admin_or_permissions(manage_events=True)
    async def excelevents(self, ctx: commands.Context):
        """Easily manage Discord Scheduled Events from Excel or pasted CSV."""
        if ctx.invoked_subcommand is None:
            await ctx.send_help(ctx.command)

    @excelevents.command(name="upload")
    async def upload(self, ctx: commands.Context):
        """
        Upload an events.xlsx file (auto-overwrites previous file).
        """
        if not ctx.message.attachments:
            await ctx.send("❌ Please attach an `.xlsx` (or `.xls`) file.")
            return

        attachment = ctx.message.attachments[0]
        if not attachment.filename.lower().endswith((".xlsx", ".xls")):
            await ctx.send("❌ Only `.xlsx` or `.xls` files are supported.")
            return

        data_path: Path = data_manager.cog_data_path(self)
        data_path.mkdir(parents=True, exist_ok=True)
        file_path = data_path / "events.xlsx"

        await attachment.save(str(file_path))
        await ctx.send(
            "✅ **File uploaded and old file replaced!**\n"
            f"Use `{ctx.prefix}excelevents check` to validate it."
        )

    @excelevents.command(name="paste")
    async def paste(self, ctx: commands.Context):
        """
        Paste raw CSV text directly.
        """
        lines = ctx.message.content.splitlines()
        csv_text = "\n".join(lines[1:]) if len(lines) > 1 else ""

        if not csv_text.strip():
            await ctx.send("❌ Please paste your CSV data after the command.")
            return

        data_path = data_manager.cog_data_path(self)
        data_path.mkdir(parents=True, exist_ok=True)
        file_path = data_path / "events.xlsx"

        try:
            reader = csv.DictReader(io.StringIO(csv_text))
            rows = list(reader)
            if not rows:
                await ctx.send("❌ No valid rows found in CSV.")
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(list(rows[0].keys()))
            for row in rows:
                ws.append(list(row.values()))
            wb.save(file_path)

            await ctx.send(
                "✅ **CSV converted and saved!**\n"
                f"Use `{ctx.prefix}excelevents check` to validate it."
            )
        except Exception as e:
            await ctx.send(f"❌ Failed to parse CSV: {type(e).__name__} – {e}")

    @excelevents.command(name="check")
    async def check(self, ctx: commands.Context):
        """Run advanced validation on the uploaded Excel/CSV file."""
        data_path = data_manager.cog_data_path(self)
        file_path = data_path / "events.xlsx"

        await ctx.send("🔍 **Running advanced validation...**")

        errors, warnings = await self._validate_excel(file_path, ctx.guild)

        if errors:
            error_text = "\n".join([f"❌ {msg}" for msg in errors])
            await ctx.send(
                f"**❌ Validation Failed – Fix these errors:**\n{error_text}\n\n"
                f"After fixing, re-upload and run `{ctx.prefix}excelevents check` again."
            )
        elif warnings:
            warn_text = "\n".join([f"⚠️ {msg}" for msg in warnings])
            await ctx.send(
                f"**✅ File is valid with warnings:**\n{warn_text}\n\n"
                f"You may proceed with `{ctx.prefix}excelevents sync`."
            )
        else:
            await ctx.send("✅ **Perfect! No errors or warnings.** Ready to sync.")

    @excelevents.command(name="sync")
    async def sync(self, ctx: commands.Context):
        """Process the uploaded Excel/CSV and sync Discord events + optional announcements."""
        if not ctx.guild.me.guild_permissions.manage_events:
            await ctx.send("❌ I need the **Manage Events** permission.")
            return

        data_path = data_manager.cog_data_path(self)
        file_path = data_path / "events.xlsx"
        if not file_path.exists():
            await ctx.send("❌ No file found. Use `upload` or `paste` first.")
            return

        errors, warnings = await self._validate_excel(file_path, ctx.guild)
        if errors:
            await ctx.send("⚠️ **Validation failed.** Run `excelevents check` first.")
            return
        if warnings:
            await ctx.send(f"⚠️ **Validation passed with warnings.**\n{chr(10).join([f'• {w}' for w in warnings])}\nContinuing...")

        await ctx.send("🔄 **Syncing events…** (real-time feedback below)")

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            headers = [
                str(cell.value).strip().lower() if cell.value is not None else ""
                for cell in next(ws.iter_rows(min_row=1, max_row=1))
            ]

            mappings = await self.config.guild(ctx.guild).event_mappings()
            new_mappings: Dict[str, int] = {}
            active_keys = set()
            processed = 0
            new_events_created: list[discord.ScheduledEvent] = []

            row_num = 1
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_num += 1
                if not row or all(v is None for v in row):
                    continue

                data = {
                    headers[i]: row[i]
                    for i in range(len(row))
                    if i < len(headers) and headers[i]
                }

                name = str(data.get("name", "")).strip()
                await ctx.send(
                    f"**Row {row_num}** → Name: `{name}` | "
                    f"Start: `{data.get('start')}` | ChannelID: `{data.get('channelid')}`"
                )

                if not name:
                    continue

                key = self._normalize_key(name)
                active_keys.add(key)

                start_time = await self._parse_datetime(data.get("start"))
                end_time = await self._parse_datetime(data.get("end"))

                if not start_time:
                    continue

                # Try to update existing event
                if key in mappings:
                    try:
                        event = await ctx.guild.fetch_scheduled_event(mappings[key])
                        await event.edit(
                            name=name,
                            description=str(data.get("description", "")).strip()[:1000] or None,
                            start_time=start_time,
                            end_time=end_time,
                        )
                        new_mappings[key] = event.id
                        processed += 1
                        await ctx.send(f"✅ **Event updated:** [{name}]({event.url})")
                        continue
                    except Exception:
                        try:
                            old_event = await ctx.guild.fetch_scheduled_event(mappings[key])
                            await old_event.delete()
                        except Exception:
                            pass

                # Create new event
                new_event = await self._create_event(ctx.guild, data)
                if new_event:
                    new_mappings[key] = new_event.id
                    new_events_created.append(new_event)
                    processed += 1
                    await ctx.send(f"✅ **Event created successfully:** [{new_event.name}]({new_event.url})")

            # Clean up removed events
            deleted = 0
            for old_key, old_id in list(mappings.items()):
                if old_key not in active_keys:
                    try:
                        event = await ctx.guild.fetch_scheduled_event(old_id)
                        await event.delete()
                        deleted += 1
                    except Exception:
                        pass

            await self.config.guild(ctx.guild).event_mappings.set(new_mappings)
            await self.config.guild(ctx.guild).last_synced.set(
                datetime.now(timezone.utc).isoformat()
            )

            # Announcement Mode (creation announcements)
            announcement_mode = await self.config.guild(ctx.guild).announcement_mode()
            announced = 0
            if announcement_mode and new_events_created:
                ann_ch_id = await self.config.guild(ctx.guild).announcement_channel()
                if ann_ch_id:
                    ann_channel = ctx.guild.get_channel(ann_ch_id)
                    if ann_channel and ann_channel.permissions_for(ctx.guild.me).send_messages:
                        for event in new_events_created:
                            try:
                                embed = self._create_event_embed(event)
                                await ann_channel.send(embed=embed)
                                announced += 1
                                await asyncio.sleep(0.8)
                            except Exception:
                                pass

            result = (
                f"**✅ FINAL RESULT**\n"
                f"• Processed: **{processed}**\n"
                f"• Active now: **{len(new_mappings)}**\n"
                f"• Deleted: **{deleted}**"
            )
            if announced:
                result += f"\n📢 **Announced {announced} new events**!"
            await ctx.send(result)

        except Exception as e:
            await ctx.send(f"❌ Sync failed: {type(e).__name__}: {e}")

    @excelevents.command(name="status")
    async def status(self, ctx: commands.Context):
        """Show full status of the cog (file, events, announcement, and reminder settings)."""
        data_path = data_manager.cog_data_path(self)
        file_path = data_path / "events.xlsx"
        mappings = await self.config.guild(ctx.guild).event_mappings()

        ann_mode = await self.config.guild(ctx.guild).announcement_mode()
        ann_ch_id = await self.config.guild(ctx.guild).announcement_channel()
        ann_ch = ctx.guild.get_channel(ann_ch_id) if ann_ch_id else None

        rem_mode = await self.config.guild(ctx.guild).reminder_mode()
        rem_ch_id = await self.config.guild(ctx.guild).reminder_channel()
        rem_ch = ctx.guild.get_channel(rem_ch_id) if rem_ch_id else None
        rem_times = await self.config.guild(ctx.guild).reminder_minutes()

        status_msg = (
            f"**ExcelEvents Status**\n"
            f"• File exists: **{file_path.exists()}**\n"
            f"• Tracked events: **{len(mappings)}**\n"
            f"• Announcement mode: **{'✅ Enabled' if ann_mode else '❌ Disabled'}**"
        )
        if ann_ch:
            status_msg += f"\n• Announcement channel: {ann_ch.mention}"
        status_msg += (
            f"\n• Reminder mode: **{'✅ Enabled' if rem_mode else '❌ Disabled'}**"
            f"\n• Reminder channel: {'Set' if rem_ch else 'Not set'}"
        )
        if rem_ch:
            status_msg += f" → {rem_ch.mention}"
        status_msg += f"\n• Reminder times: **{rem_times}** minutes before start"
        await ctx.send(status_msg)

    @excelevents.group(name="announcement", invoke_without_command=True)
    @commands.guild_only()
    @commands.admin_or_permissions(manage_events=True)
    async def announcement_group(self, ctx: commands.Context):
        """Toggle announcement mode and set the channel for new-event embeds."""
        await ctx.send_help(ctx.command)

    @announcement_group.command(name="toggle")
    async def toggle_announcement(self, ctx: commands.Context, channel: Optional[discord.TextChannel] = None):
        """Toggle announcement mode for new/updated events."""
        config = self.config.guild(ctx.guild)

        if channel is None:
            current = await config.announcement_mode()
            new_mode = not current
            await config.announcement_mode.set(new_mode)
            await ctx.send(f"✅ Announcement mode **{'enabled' if new_mode else 'disabled'}**.")
            return

        await config.announcement_channel.set(channel.id)
        await config.announcement_mode.set(True)
        await ctx.send(
            f"✅ **Announcement mode enabled!**\n"
            f"Rich embeds will be posted in {channel.mention} when new events are created via `sync`."
        )

    @excelevents.group(name="reminder", invoke_without_command=True)
    @commands.guild_only()
    @commands.admin_or_permissions(manage_events=True)
    async def reminder_group(self, ctx: commands.Context):
        """Manage reminder settings (pre-event notifications in a channel)."""
        await ctx.send_help(ctx.command)

    @reminder_group.command(name="toggle")
    async def toggle_reminder(self, ctx: commands.Context, channel: Optional[discord.TextChannel] = None):
        """Toggle reminder mode. Provide a channel to enable + set it."""
        config = self.config.guild(ctx.guild)

        if channel is None:
            current = await config.reminder_mode()
            new_mode = not current
            await config.reminder_mode.set(new_mode)
            await ctx.send(f"✅ Reminder mode **{'enabled' if new_mode else 'disabled'}**.")
            return

        await config.reminder_channel.set(channel.id)
        await config.reminder_mode.set(True)
        await ctx.send(
            f"✅ **Reminder mode enabled!**\n"
            f"Pre-event reminders will be automatically posted in {channel.mention} "
            f"at the configured times before each event starts."
        )

    @reminder_group.command(name="times")
    async def reminder_times(self, ctx: commands.Context, *minutes: int):
        """Set reminder times in minutes before the event (example: 60 15 5)."""
        config = self.config.guild(ctx.guild)

        if not minutes:
            current = await config.reminder_minutes()
            await ctx.send(f"Current reminder times: **{current}** minutes before start.")
            return

        valid = [m for m in minutes if m > 0]
        if not valid:
            await ctx.send("❌ Please provide positive numbers (minutes before start).")
            return

        await config.reminder_minutes.set(valid)
        await ctx.send(f"✅ Reminder times updated to: **{valid}** minutes before events.")

    @excelevents.command(name="clear")
    async def clear(self, ctx: commands.Context):
        """Delete the events.xlsx file and reset all tracked mappings."""
        data_path = data_manager.cog_data_path(self)
        file_path = data_path / "events.xlsx"

        if file_path.exists():
            file_path.unlink()
            await self.config.guild(ctx.guild).event_mappings.set({})
            await ctx.send("✅ **Events file deleted** and mappings reset.")
        else:
            await ctx.send("No file to clear.")

    def cog_unload(self):
        """Clean shutdown – cancel reminder task."""
        if self.reminder_task and not self.reminder_task.done():
            self.reminder_task.cancel()
        pass
